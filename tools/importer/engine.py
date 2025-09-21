# tools/importer/engine.py
import os
import csv

# Excel reading via openpyxl
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

def _normalize_key(k: str):
    """Normalize a key string: strip and lower."""
    return k.strip() if isinstance(k, str) else k

def read_key_value_file(path: str):
    """
    Read a file containing rows of key | value and parse keys of the form:
        table>field
    Returns a dict of the form:
        { 'client': {'sds_id': '43468172', 'entity_name': 'Amazon'}, 
          'merchant': {'merchant_name': 'M1', 'client_sds_id': '43468172'},
          'ratesheet': { ... } }
    Raises FileNotFoundError, ValueError (unsupported extension), or RuntimeError if openpyxl missing.
    """
    path = os.path.abspath(path)
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")

    ext = os.path.splitext(path)[1].lower()
    parsed = {}

    def add_key_value(raw_key, raw_val):
        if raw_key is None:
            return
        key = str(raw_key).strip()
        # ignore blank keys
        if not key:
            return
        # support either "table>field" or just "field" (treat as unknown -> top-level)
        if '>' in key:
            table, field = key.split('>', 1)
            table = table.strip().lower()
            field = field.strip().lower()
        else:
            table = None
            field = key.strip().lower()
        val = '' if raw_val is None else str(raw_val).strip()
        if table:
            parsed.setdefault(table, {})[field] = val
        else:
            # fallback: put in a special dictionary
            parsed.setdefault('_unknown', {})[field] = val

    if ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        if load_workbook is None:
            raise RuntimeError("openpyxl is required to read Excel files: pip install openpyxl")
        wb = load_workbook(filename=path, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            # first two columns only
            if len(row) >= 2:
                raw_key, raw_val = row[0], row[1]
                if raw_key is None:
                    continue
                add_key_value(raw_key, raw_val)
            elif len(row) == 1 and row[0] is not None:
                # single-column row - ignore
                continue
    elif ext == '.csv':
        with open(path, newline='', encoding='utf-8') as fh:
            reader = csv.reader(fh)
            for row in reader:
                if not row:
                    continue
                if len(row) >= 2:
                    add_key_value(row[0], row[1])
    else:
        raise ValueError("Unsupported file extension. Use .xlsx or .csv")

    return parsed

# -------------------------
# Insert helpers (DB-aware)
# -------------------------
def insert_from_parsed(db, parsed: dict):
    """
    Insert content from parsed dict into DB.
    Rules:
    - If 'client' data present, insert client first (requires sds_id & entity_name).
    - If 'merchant' present, requires client_sds_id (or will use inserted client)
    - If 'ratesheet' present, requires client_sds_id (or will use inserted client). If merchant specified by name,
      we will try to find merchant for that client, or create it if missing.
    Returns a summary dict of inserted ids, e.g. {'client': 43468172, 'merchant': 7, 'ratesheet': 3}
    Raises ValueError on validation errors.
    """
    summary = {}
    # Utility: safe getter
    def g(table, field, default=None):
        return parsed.get(table, {}).get(field, default)

    # 1) Insert client (if present)
    client_id = None
    if 'client' in parsed:
        client_data = parsed['client']
        if 'sds_id' not in client_data or not client_data.get('entity_name'):
            raise ValueError("Client data must include 'client>sds_id' and 'client>entity_name'")
        try:
            client_id = int(client_data['sds_id'])
        except Exception:
            # allow non-int? enforce int
            raise ValueError("client>sds_id must be an integer value")
        # call insert_client (uses INSERT OR IGNORE semantics in DBManager)
        bank = client_data.get('bank_user_id') or None
        tz = client_data.get('timezone') or None
        eod = client_data.get('end_of_day') or None
        # insert_client returns lastrowid for an inserted row; if row existed insert_client may do IGNORE and return 0
        new = db.insert_client(client_id, client_data['entity_name'], bank, tz, eod)
        # our clients use the sds_id as primary key; return the sds_id as identifier
        summary['client'] = client_id
    # 2) Insert merchant (if present)
    merchant_id = None
    if 'merchant' in parsed:
        m = parsed['merchant']
        # ensure we have a client_sds_id either in merchant data or from inserted client
        client_sds = None
        if m.get('client_sds_id'):
            try:
                client_sds = int(m.get('client_sds_id'))
            except Exception:
                raise ValueError("merchant>client_sds_id must be integer")
        elif client_id is not None:
            client_sds = client_id
        else:
            raise ValueError("Merchant requires merchant>client_sds_id or a client block in file")

        if 'merchant_name' not in m or not m.get('merchant_name'):
            raise ValueError("Merchant data must include 'merchant>merchant_name'")

        code = m.get('merchant_code') or None
        merchant_id = db.insert_merchant(client_sds, m['merchant_name'], code)
        summary['merchant'] = merchant_id

    # 3) Insert ratesheet (if present)
    ratesheet_id = None
    if 'ratesheet' in parsed:
        r = parsed['ratesheet']
        # find client_sds
        client_sds = None
        if r.get('client_sds_id'):
            try:
                client_sds = int(r.get('client_sds_id'))
            except Exception:
                raise ValueError("ratesheet>client_sds_id must be integer")
        elif client_id is not None:
            client_sds = client_id
        else:
            raise ValueError("Ratesheet requires ratesheet>client_sds_id or a client block in file")

        # merchant resolution: can be merchant_id or merchant_name
        merchant_ref_id = None
        if r.get('merchant_id'):
            try:
                merchant_ref_id = int(r.get('merchant_id'))
            except Exception:
                raise ValueError("ratesheet>merchant_id must be integer")
        elif r.get('merchant_name'):
            mname = r.get('merchant_name')
            # find merchant by name for the client
            found = None
            ml = db.fetch_merchants_by_client(client_sds)
            for mm in ml:
                if mm.get('merchant_name', '').strip().lower() == mname.strip().lower():
                    found = mm
                    break
            if found:
                merchant_ref_id = found['merchant_id']
            else:
                # create merchant automatically if not found
                merchant_ref_id = db.insert_merchant(client_sds, mname)
        # now prepare ratesheet fields
        eff = r.get('effective_date') or None
        exp = r.get('expiry_date') or None
        details = r.get('rate_details') or None
        ratesheet_id = db.insert_ratesheet(client_sds, merchant_ref_id, eff, exp, details)
        summary['ratesheet'] = ratesheet_id

    # 4) If file had unknown top-level keys under '_unknown', return them in summary too
    if '_unknown' in parsed:
        summary['_unknown'] = parsed['_unknown']

    if not summary:
        # nothing to insert
        raise ValueError("No recognized table>field keys found. Provide keys like 'client>sds_id' or 'merchant>merchant_name' or 'ratesheet>effective_date'.")

    return summary
